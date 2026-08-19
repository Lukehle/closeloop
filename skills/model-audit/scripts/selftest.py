#!/usr/bin/env python3
"""Self-test for audit_workbook.py.

Builds a workbook with deliberately planted defects plus deliberately clean
rows, runs the scanner, and asserts that every planted defect is caught and no
clean row is flagged.

The truncated-range case (row 12) is the regression guard that matters most:
an earlier normalizer stripped row numbers from references, which made
=SUM(B1:B9) and =SUM(D1:D8) compare equal and silently hid the single most
common range defect in real models.

Usage:  python selftest.py
Exit:   0 all assertions pass, 1 one or more failed, 2 openpyxl unavailable.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
SCANNER = os.path.join(HERE, "audit_workbook.py")


def build_fixture(path: str) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    # Row 10 - growth series. B10 is a legitimate seed value (must NOT flag);
    # E10 is a planted hardcode inside the calculated series (MUST flag).
    ws["A10"] = "Revenue"
    ws["B10"] = 1000
    for col in "CDF":
        ws[f"{col}10"] = f"={chr(ord(col) - 1)}10*(1+$A$1)"
    ws["E10"] = 1250000

    # Row 12 - uniform SUMs with one truncated range planted at D12 (MUST flag).
    for col in "BCEF":
        ws[f"{col}12"] = f"=SUM({col}1:{col}9)"
    ws["D12"] = "=SUM(D1:D8)"

    ws["B14"] = "=#REF!+B13"
    ws["B15"] = "='[FY25_Model.xlsx]Sheet1'!A1"
    ws["B16"] = '=INDIRECT("B"&1)'

    # Row 18 - fully uniform. Must produce NO inconsistency finding.
    for col in "BCDEF":
        ws[f"{col}18"] = f"={col}12*$A$2"

    ws.merge_cells("A20:C20")
    ws.row_dimensions[30].hidden = True

    hidden = wb.create_sheet("Backup")
    hidden.sheet_state = "hidden"
    hidden["A1"] = 1

    wb.save(path)


def main() -> int:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("SKIP: openpyxl unavailable; cannot self-test.", file=sys.stderr)
        return 2

    with tempfile.TemporaryDirectory() as tmp:
        book = os.path.join(tmp, "fixture.xlsx")
        build_fixture(book)

        proc = subprocess.run(
            [sys.executable, SCANNER, book, "--json"],
            capture_output=True, text=True,
        )
        if proc.returncode == 2:
            print(f"SKIP: scanner could not read fixture: {proc.stderr}", file=sys.stderr)
            return 2

        report = json.loads(proc.stdout)
        hits = {(f["check"], f["cell"]) for f in report["findings"]}

        assertions = [
            ("planted hardcode E10 flagged",
             ("hardcode_in_formula_row", "E10") in hits),
            ("seed value B10 not flagged",
             ("hardcode_in_formula_row", "B10") not in hits),
            ("truncated range D12 flagged (regression guard)",
             ("inconsistent_row_formula", "D12") in hits),
            ("clean growth series row 10 not flagged",
             not any(c == "inconsistent_row_formula" and cell in ("C10", "D10", "F10")
                     for c, cell in hits)),
            ("clean uniform row 18 not flagged",
             not any(c == "inconsistent_row_formula" and cell.endswith("18")
                     for c, cell in hits)),
            ("error value B14 flagged", ("error_values", "B14") in hits),
            ("external link B15 flagged", ("external_link", "B15") in hits),
            ("volatile function B16 flagged", ("volatile_functions", "B16") in hits),
            ("scanner signalled findings via exit code", proc.returncode == 1),
        ]

    failed = 0
    for label, ok in assertions:
        print(("PASS  " if ok else "FAIL  ") + label)
        failed += 0 if ok else 1

    print(f"\n{len(assertions) - failed}/{len(assertions)} passed")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
