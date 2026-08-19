#!/usr/bin/env python3
"""Deterministic spreadsheet integrity scanner for the closeloop model-audit skill.

Reads an .xlsx workbook and reports mechanically detectable defects: hardcoded
constants inside formula rows, structurally inconsistent formulas across a row,
error values, external links, volatile functions, merged cells inside data
regions, hidden content, and deeply nested formulas.

It reports what it can prove from the file. It makes no judgement about whether
the model is financially correct - that is the reasoned pass in SKILL.md.

Usage:
    python audit_workbook.py MODEL.xlsx [--json] [--min-severity high]

Exit codes:
    0  no findings
    1  findings present
    2  could not read the workbook (missing file, missing dependency, bad format)

Dependency: openpyxl. If unavailable, exits 2 with the degraded-mode pointer
rather than pretending to have scanned anything.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict

ERROR_VALUES = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!", "#NULL!")
VOLATILE = ("INDIRECT", "OFFSET", "TODAY", "NOW", "RAND", "RANDBETWEEN")

SEVERITY_ORDER = {"critical": 0, "high": 1, "medium": 2, "low": 3}

# A formula row needs at least this many formula cells before a lone constant
# in it is suspicious. Below this the row is probably not a calculated series.
MIN_FORMULAS_FOR_HARDCODE_CHECK = 3

# Nesting depth beyond which a formula stops being reviewable by a human.
MAX_NESTING_DEPTH = 6


_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _norm_formula(formula: str, coord: str) -> str:
    """Reduce a formula to its structural skeleton, R1C1-style.

    Cell references become offsets *relative to the cell containing them*, so
    two cells implementing the same calculation in adjacent columns compare
    equal, while a genuine structural difference stands out.

    This relative encoding is what makes the classic truncated-range defect
    visible: =SUM(B1:B9) in B12 and =SUM(D1:D8) in D12 look identical if you
    merely strip the references, but as offsets they are R[-11]C[0]:R[-3]C[0]
    versus R[-11]C[0]:R[-4]C[0] - different, and correctly flagged.

    Absolute references ($A$1) keep their literal address, because an absolute
    reference is meant to be the same cell everywhere in the row.
    """
    m = _REF_RE.fullmatch(coord.upper().replace("$", ""))
    if m:
        base_col, base_row = _col_to_num(m.group(2)), int(m.group(4))
    else:  # merged range or other non-cell coordinate; fall back to absolute
        base_col = base_row = None

    def repl(match: re.Match) -> str:
        col_abs, col, row_abs, row = match.groups()
        col_n, row_n = _col_to_num(col), int(row)
        if base_col is None:
            return f"A[{col_n}]R[{row_n}]"
        c = f"C[{col_n}]" if col_abs else f"C[{col_n - base_col:+d}]"
        r = f"R[{row_n}]" if row_abs else f"R[{row_n - base_row:+d}]"
        return r + c

    f = formula.upper()
    f = re.sub(r'"[^"]*"', "STR", f)          # string literals first
    f = _REF_RE.sub(repl, f)                  # references -> relative offsets

    # Collapse remaining numeric literals, but never touch digits inside the
    # [...] offset tokens just emitted - those offsets ARE the structure, and
    # eating them makes a truncated range indistinguishable from a full one.
    f = re.sub(r"\[[^\]]*\]|\d+(?:\.\d+)?",
               lambda m: m.group(0) if m.group(0).startswith("[") else "N", f)
    f = re.sub(r"\s+", "", f)
    return f


def _nesting_depth(formula: str) -> int:
    depth = max_depth = 0
    for ch in formula:
        if ch == "(":
            depth += 1
            max_depth = max(max_depth, depth)
        elif ch == ")":
            depth = max(0, depth - 1)
    return max_depth


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _is_numeric_constant(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _finding(sev, check, sheet, cell, detail, formula=None):
    out = {
        "severity": sev,
        "check": check,
        "sheet": sheet,
        "cell": cell,
        "detail": detail,
    }
    if formula:
        out["formula"] = formula[:200]
    return out


def scan_sheet(ws) -> list[dict]:
    findings: list[dict] = []
    name = ws.title

    rows_formulas: dict[int, list] = defaultdict(list)
    rows_constants: dict[int, list] = defaultdict(list)

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue

            if _is_formula(v):
                rows_formulas[cell.row].append((cell.coordinate, v))

                for err in ERROR_VALUES:
                    if err in v:
                        findings.append(_finding(
                            "critical", "error_values", name, cell.coordinate,
                            f"formula contains {err}", v))
                        break

                # External workbook link: [Book.xlsx] or a UNC/absolute path.
                if "[" in v and "]" in v:
                    findings.append(_finding(
                        "high", "external_link", name, cell.coordinate,
                        "references another workbook", v))

                for fn in VOLATILE:
                    if re.search(rf"\b{fn}\s*\(", v.upper()):
                        findings.append(_finding(
                            "medium", "volatile_functions", name, cell.coordinate,
                            f"uses volatile function {fn}", v))
                        break

                depth = _nesting_depth(v)
                if depth > MAX_NESTING_DEPTH:
                    findings.append(_finding(
                        "medium", "deep_nesting", name, cell.coordinate,
                        f"nesting depth {depth} exceeds {MAX_NESTING_DEPTH}", v))

            elif _is_numeric_constant(v):
                rows_constants[cell.row].append((cell.coordinate, v))

            elif isinstance(v, str):
                for err in ERROR_VALUES:
                    if v.strip() == err:
                        findings.append(_finding(
                            "critical", "error_values", name, cell.coordinate,
                            f"cached error value {err}"))
                        break

    # Hardcoded constants sitting inside an otherwise-calculated row.
    #
    # The leftmost populated cell of a row is exempt: a seed/opening value at the
    # start of a calculated series is conventional and correct, and flagging it
    # produces noise that buries the real findings further right.
    for row_idx, formulas in rows_formulas.items():
        if len(formulas) < MIN_FORMULAS_FOR_HARDCODE_CHECK:
            continue
        constants = rows_constants.get(row_idx, [])
        if not constants:
            continue
        populated = [c for c, _ in constants] + [c for c, _ in formulas]
        leftmost = min(populated, key=lambda c: _col_to_num(
            _REF_RE.fullmatch(c).group(2)))
        for coord, const in constants:
            if coord == leftmost:
                continue  # seed value at the head of the series
            findings.append(_finding(
                "critical", "hardcode_in_formula_row", name, coord,
                f"constant {const!r} in a row with {len(formulas)} formula cells "
                f"- this value will not respond to assumption changes"))

    # Structurally inconsistent formulas across a row.
    for row_idx, formulas in rows_formulas.items():
        if len(formulas) < MIN_FORMULAS_FOR_HARDCODE_CHECK:
            continue
        shapes: dict[str, list[str]] = defaultdict(list)
        for coord, f in formulas:
            shapes[_norm_formula(f, coord)].append(coord)
        if len(shapes) > 1:
            ranked = sorted(shapes.items(), key=lambda kv: len(kv[1]), reverse=True)
            majority_n = len(ranked[0][1])
            for shape, coords in ranked[1:]:
                # Only flag genuine minorities, not a two-way split.
                if len(coords) < majority_n:
                    for coord in coords:
                        formula = next(f for c, f in formulas if c == coord)
                        findings.append(_finding(
                            "high", "inconsistent_row_formula", name, coord,
                            f"formula structure differs from {majority_n} other "
                            f"cells in row {row_idx}", formula))

    if ws.merged_cells.ranges:
        for rng in list(ws.merged_cells.ranges)[:20]:
            findings.append(_finding(
                "medium", "merged_cells", name, str(rng),
                "merged range - breaks references, sorting, and structured reads"))

    hidden_rows = [i for i, d in ws.row_dimensions.items() if d.hidden]
    hidden_cols = [k for k, d in ws.column_dimensions.items() if d.hidden]
    if hidden_rows:
        findings.append(_finding(
            "medium", "hidden_content", name, "-",
            f"{len(hidden_rows)} hidden row(s): {hidden_rows[:15]}"))
    if hidden_cols:
        findings.append(_finding(
            "medium", "hidden_content", name, "-",
            f"{len(hidden_cols)} hidden column(s): {hidden_cols[:15]}"))

    if not ws.protection.sheet and rows_formulas:
        findings.append(_finding(
            "medium", "no_protection", name, "-",
            "sheet has formulas but no protection - inputs and calculations are "
            "equally overwritable"))

    return findings


def main() -> int:
    ap = argparse.ArgumentParser(description="closeloop model-audit mechanical scanner")
    ap.add_argument("workbook", help="path to an .xlsx file")
    ap.add_argument("--json", action="store_true", help="emit JSON only")
    ap.add_argument("--min-severity", default="low",
                    choices=["critical", "high", "medium", "low"],
                    help="suppress findings below this severity")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print(
            "ERROR: openpyxl is not available, so nothing was scanned.\n"
            "Install it (pip install openpyxl) or follow the Degraded mode section\n"
            "of the model-audit SKILL.md, which covers the same defect taxonomy by hand.",
            file=sys.stderr)
        return 2

    try:
        wb = openpyxl.load_workbook(args.workbook, data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001 - surface any load failure verbatim
        print(f"ERROR: could not read {args.workbook}: {exc}", file=sys.stderr)
        return 2

    findings: list[dict] = []
    formula_cells = 0

    for ws in wb.worksheets:
        findings.extend(scan_sheet(ws))
        for row in ws.iter_rows():
            for cell in row:
                if _is_formula(cell.value):
                    formula_cells += 1

    hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    if hidden_sheets:
        findings.append(_finding(
            "medium", "hidden_content", "-", "-",
            f"hidden sheet(s): {hidden_sheets}"))

    floor = SEVERITY_ORDER[args.min_severity]
    findings = [f for f in findings if SEVERITY_ORDER[f["severity"]] <= floor]
    findings.sort(key=lambda f: (SEVERITY_ORDER[f["severity"]], f["sheet"], f["cell"]))

    counts = defaultdict(int)
    for f in findings:
        counts[f["severity"]] += 1

    report = {
        "workbook": args.workbook,
        "sheets": len(wb.worksheets),
        "formula_cells": formula_cells,
        "counts": dict(counts),
        "findings": findings,
    }

    if args.json:
        print(json.dumps(report, indent=2, default=str))
    else:
        print(json.dumps(report, indent=2, default=str))
        print(
            f"\n# {len(findings)} finding(s): "
            f"{counts['critical']} critical, {counts['high']} high, "
            f"{counts['medium']} medium, {counts['low']} low",
            file=sys.stderr)
        print("# Mechanical scan only. Run the reasoned review in SKILL.md before "
              "trusting this model.", file=sys.stderr)

    return 1 if findings else 0


if __name__ == "__main__":
    sys.exit(main())
